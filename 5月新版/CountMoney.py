# -*- coding: UTF-8 -*-

'''
__author__="zf"
__mtime__ = '2016/11/8/21/38'
__des__: 简单的读取文件
__lastchange__:'2016/11/16'
'''
from __future__ import division
import xlrd
import os
import math
from xlwt import Workbook, Formula
import xlrd
import sys
import types
import copy  
import time  
import pandas
from openpyxl import Workbook
import openpyxl,pprint
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
from xlrd import xldate_as_tuple
def is_chinese(uchar): 
        """判断一个unicode是否是汉字"""
        if uchar >= u'/u4e00' and uchar<=u'/u9fa5':
                return True
        else:
                return False

                
def is_num(unum):
	try:
		unum+1
	except TypeError:
		return 0
	else:
		return 1

#不带颜色的读取
def filename(content):
	#打开文件,如果是非xlsx文件转换之
	global workbook,file_excel
	file_excel=str(content)

	file=(file_excel+'.xlsx').decode('utf-8')#文件名及中文合理性
	if not os.path.exists(file):#判断文件是否存在

		x = pd.read_excel(file_excel+'.xls')
		x.to_excel(file_excel+'.xlsx', index=False)
		
		
		if not os.path.exists(file):
			print ("文件不存在")

	wb = openpyxl.load_workbook(file_excel+'.xlsx')
	print (wb,'suicce')
	return wb

def readexcel(content):
	wb=filename(content)
	
	sheet=wb.get_active_sheet()
	UserData={}
	print (sheet.rows,'111111')

	#遍历第一行，获取所有人名并创建字典,格式{1xxx:{},2xxxx{}}
	NeedDealLie=[]

	for cell in list(sheet.rows)[1]:
		a=cell.value
		try:
			if  u'.' in  a:
				#人名行的开始位置
				NeedDealLie.append(cell.coordinate)

				if u'/' in a:
					NameTypeList=a.split(u'.')[1].split(u'/')
					Gongxu=a.split(u'.')[0]
					for NameType in NameTypeList:
						NameType=Gongxu+u'.'+NameType
						UserData.setdefault(NameType, {})
				else:
					NameType=a
					UserData.setdefault(NameType, {})
			else:
				#第四列的开始行位置
				if a==u'塑料袋':
					DiSiSpec=cell.coordinate
		except:
			pass
	for key in UserData:
		print key
	print UserData
	print NeedDealLie,DiSiSpec
	# for row in range(2,max_row+1):

	#遍历所有，给字典加上值,格式{1xxx:{1：{日期：xxx，数量：xxx，型号：xxxx}},2xxxx{}}

	for row in range(3, sheet.max_row + 1):
		KeRenBianHao=sheet['A' + str(row)].value
		ZongChengXingHao= sheet['B' + str(row)].value
		

		for x in NeedDealLie:
			#x[0]是支取坐标的Y坐标
			if sheet[x[0]+str(row)].value !=None:
				NeedPlus=sheet[x].value
				#获得数量行，前一行日期行的y坐标
				RiqiHang=get_column_letter(column_index_from_string(x[0])-1)
				Riqi=sheet[RiqiHang+str(row)].value
				ShuLiang= sheet[x[0]+str(row)].value




				for key in UserData:
					if key.split(u'.')[1]  in  NeedPlus:

						UserData[key].setdefault(row, {'KeRenBianHao':0,'ZongChengXingHao':0,'ShuLiang':0,'Riqi':0})
						UserData[key][row]['KeRenBianHao']=KeRenBianHao
						UserData[key][row]['ZongChengXingHao']=ZongChengXingHao
						UserData[key][row]['ShuLiang']=ShuLiang
						UserData[key][row]['Riqi']=Riqi

						if key.split(u'.')[0]==u'4':
							
							#第四道的获取后三行的信息
							SuLiaoDai=sheet[DiSiSpec[0]+str(row)].value
							BiaoTieHang=get_column_letter(column_index_from_string(DiSiSpec[0])+1)
							NeiHeHang=get_column_letter(column_index_from_string(DiSiSpec[0])+2)
							XiaoBiaoTie=sheet[BiaoTieHang+str(row)].value
							XiaoNeiHe=sheet[BiaoTieHang+str(row)].value



							UserData[key][row]['SuLiaoDai']=SuLiaoDai
							UserData[key][row]['XiaoBiaoTie']=XiaoBiaoTie
							UserData[key][row]['XiaoNeiHe']=XiaoNeiHe
	print('Writing results...')
	resultFile = open('census2010.py', 'w')
	resultFile.write('allData = ' + pprint.pformat(UserData))
	resultFile.close()
	print('Done.')
	return UserData




def readjiage(content):
	wb=filename(content)
	money=[]
	sheetSum=wb.get_sheet_names()

	for x in sheetSum:
		BiaoXinXi=[]
		sheet = wb.get_sheet_by_name(x)
		Gongxu=sheet.title
		BiaoXinXi.append(Gongxu)
		for row in sheet.rows:
			mid=[]
			for cell in row:		
				if cell.value!=None:
					mid.append(cell.value)
			if mid!=[]:
				BiaoXinXi.append(mid)
		money.append(BiaoXinXi)

	# 	print BiaoXinXi
	# # return money
	# print money,'montytyyt'
	book = Workbook()
	sheet1 = book.add_sheet('Sheet 1')
	for i in range(len(money)):
		for j in range (len(money[i])):
			for z in range (len(money[i][j]))
				if is_chinese(money[i][j]):
					money[i][j].encode('utf-8')
				# elif not money[i] and money[i]!=0: 
				# 	print "空值",
				elif is_num(money[i][j])==1:
					if math.modf(money[i][j])[0]==0 or money[i][j]==0:#获取数字的整数和小数
						money[i][j]=int(money[i][j])#将浮点数化成整数
				sheet1.write(i,j,money[i][j])
	book.save('allmesg.xls')#存储excel
	book = xlrd.open_workbook('allmesg.xls')

if __name__ == "__main__":

	# readexcel1('newmy')
	readexcel('12345')
	readjiage('newmy')
	# readexcel2('taizhang')
	